"""
Builds the downloadable attendance .xlsx sheet, matching the report's
requirement that attendance is "stored in excel format".
"""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import config
import db


def build_attendance_workbook(date_filter=None):
    """
    Build an .xlsx workbook of attendance records.
    If date_filter is given (YYYY-MM-DD), only that date's records are included;
    otherwise every recorded date is included, one row per (student, date, lecture).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["Roll No", "Name", "Date", "Lecture", "Status", "Marked At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    rows = db.get_attendance_for_date(date_filter) if date_filter else db.get_all_attendance()

    for row in rows:
        ws.append([row["roll_no"], row["name"], row["date"], row["lecture"], row["status"], row["marked_at"]])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)

    filename = f"attendance_{date_filter or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(config.EXPORTS_DIR, filename)
    wb.save(filepath)
    return filepath
